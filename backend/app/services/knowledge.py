"""
Knowledge Base Service
───────────────────────
Ingests documents (PDF, URL, Excel, YouTube, manual)
and stores embeddings in pgvector for RAG retrieval.
"""
import os
import io
import json
import httpx
import asyncio
from typing import Optional
from loguru import logger
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, text, delete

from app.core.config import settings
from app.models.knowledge import KnowledgeSource, KnowledgeChunk


# ─── EMBEDDING ──────────────────────────────────────────────
async def embed_text(text_content: str) -> list[float]:
    """Generate embedding using OpenAI text-embedding-3-small"""
    import openai
    client = openai.AsyncOpenAI(api_key=settings.OPENAI_API_KEY)
    response = await client.embeddings.create(
        model="text-embedding-3-small",
        input=text_content[:8000],  # Limit chunk size
    )
    return response.data[0].embedding


async def embed_text_batch(texts: list[str]) -> list[list[float]]:
    import openai
    client = openai.AsyncOpenAI(api_key=settings.OPENAI_API_KEY)
    # Process in batches of 100
    all_embeddings = []
    for i in range(0, len(texts), 100):
        batch = texts[i:i+100]
        response = await client.embeddings.create(
            model="text-embedding-3-small",
            input=batch,
        )
        all_embeddings.extend([d.embedding for d in response.data])
    return all_embeddings


# ─── SEARCH ─────────────────────────────────────────────────
async def search_knowledge(
    query: str,
    tenant_id: str,
    db: AsyncSession,
    top_k: int = 5,
) -> str:
    """Search knowledge base using vector similarity"""
    try:
        # Get query embedding
        query_embedding = await embed_text(query)
        embedding_str = "[" + ",".join(str(x) for x in query_embedding) + "]"

        # Vector similarity search
        result = await db.execute(
            text("""
                SELECT content, metadata,
                       1 - (embedding <=> :embedding::vector) AS similarity
                FROM knowledge_chunks
                WHERE tenant_id = :tenant_id
                ORDER BY embedding <=> :embedding::vector
                LIMIT :top_k
            """),
            {
                "embedding": embedding_str,
                "tenant_id": tenant_id,
                "top_k": top_k,
            }
        )
        rows = result.fetchall()

        if not rows:
            return ""

        chunks = []
        for row in rows:
            if row.similarity > 0.5:  # Only include relevant chunks
                chunks.append(f"[Relevance: {row.similarity:.0%}]\n{row.content}")

        return "\n\n---\n\n".join(chunks)

    except Exception as e:
        logger.error(f"Knowledge search error: {e}")
        return ""


# ─── INGESTION ──────────────────────────────────────────────
async def ingest_source(source_id: str, db: AsyncSession):
    """Main ingestion dispatcher"""
    result = await db.execute(
        select(KnowledgeSource).where(KnowledgeSource.id == source_id)
    )
    source = result.scalar_one_or_none()
    if not source:
        return

    source.status = "processing"
    await db.commit()

    try:
        if source.type == "website":
            # Default to depth 1 (recursive) for website sources
            chunks = await _ingest_website(source.url, depth=1)
        elif source.type == "pdf":
            chunks = await _ingest_pdf(source.file_path)
        elif source.type == "excel":
            chunks = await _ingest_excel(source.file_path)
        elif source.type == "youtube":
            chunks = await _ingest_youtube(source.url)
        elif source.type == "manual":
            chunks = [source.url]  # url field stores manual content
        else:
            raise ValueError(f"Unknown source type: {source.type}")

        # Remove old chunks for this source
        await db.execute(
            delete(KnowledgeChunk).where(KnowledgeChunk.source_id == source_id)
        )

        # Embed and store chunks
        if chunks:
            embeddings = await embed_text_batch(chunks)
            for chunk_text, embedding in zip(chunks, embeddings):
                chunk = KnowledgeChunk(
                    tenant_id=source.tenant_id,
                    source_id=source.id,
                    content=chunk_text,
                    embedding=embedding,
                )
                db.add(chunk)

        source.status = "indexed"
        source.chunk_count = len(chunks)
        await db.commit()
        logger.info(f"Indexed {len(chunks)} chunks for source {source_id}")

    except Exception as e:
        logger.error(f"Ingestion error for source {source_id}: {e}")
        source.status = "failed"
        source.error_message = str(e)
        await db.commit()


async def _ingest_website(url: str, depth: int = 1) -> list[str]:
    """Crawl website recursively to a certain depth."""
    if depth <= 0:
        return []
    
    all_chunks = []
    visited = set()
    to_visit = [(url, 0)]

    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
        while to_visit:
            current_url, current_depth = to_visit.pop(0)
            if current_url in visited or current_depth > depth:
                continue
            
            visited.add(current_url)
            logger.info(f"Crawling {current_url} at depth {current_depth}")

            try:
                response = await client.get(current_url)
                if response.status_code != 200:
                    continue
                
                from bs4 import BeautifulSoup
                soup = BeautifulSoup(response.text, "html.parser")
                
                # Extract text
                for tag in soup(["script", "style", "nav", "footer", "header"]):
                    tag.decompose()
                all_chunks.extend(_chunk_text(soup.get_text(separator="\n", strip=True), chunk_size=500))

                # Find links for next depth
                if current_depth < depth:
                    from urllib.parse import urljoin, urlparse
                    domain = urlparse(url).netloc
                    for link in soup.find_all("a", href=True):
                        full_link = urljoin(current_url, link["href"])
                        if urlparse(full_link).netloc == domain:
                            # Keep it within same domain
                            to_visit.append((full_link, current_depth + 1))
            except Exception as e:
                logger.error(f"Failed to crawl {current_url}: {e}")

    return all_chunks


async def _ingest_pdf(file_path: str) -> list[str]:
    """Extract text from PDF"""
    import PyPDF2
    chunks = []
    with open(file_path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            text = page.extract_text()
            if text.strip():
                chunks.extend(_chunk_text(text, chunk_size=400))
    return chunks


async def _ingest_excel(file_path: str) -> list[str]:
    """Convert Excel/CSV to text chunks"""
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    chunks = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_text = ", ".join(
                f"{h}: {v}" for h, v in zip(headers, row) if v is not None
            )
            if row_text.strip():
                chunks.append(row_text)
    return chunks


async def _ingest_youtube(url: str) -> list[str]:
    """Extract transcript from YouTube video"""
    from youtube_transcript_api import YouTubeTranscriptApi
    import re
    video_id = re.search(r"(?:v=|youtu\.be/)([^&\n?#]+)", url)
    if not video_id:
        raise ValueError("Invalid YouTube URL")

    transcript = YouTubeTranscriptApi.get_transcript(video_id.group(1))
    full_text = " ".join([t["text"] for t in transcript])
    return _chunk_text(full_text, chunk_size=500)


def _chunk_text(text: str, chunk_size: int = 400, overlap: int = 50) -> list[str]:
    """Split text into overlapping chunks"""
    words = text.split()
    chunks = []
    for i in range(0, len(words), chunk_size - overlap):
        # Explicit slice to satisfy strict type checker
        chunk_words = []
        for j in range(i, min(i + chunk_size, len(words))):
            chunk_words.append(words[j])
        chunk = " ".join(chunk_words)
        if len(chunk.strip()) > 50:  # Skip tiny chunks
            chunks.append(chunk.strip())
    return chunks
